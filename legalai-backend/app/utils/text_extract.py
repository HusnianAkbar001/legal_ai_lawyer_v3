import os, re, requests, json as jsonlib, csv as csvlib
from bs4 import BeautifulSoup
from pypdf import PdfReader
from docx import Document

# Optional heavy deps (enabled by requirements update)
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

try:
    from PIL import Image
except Exception:
    Image = None

try:
    import pytesseract
except Exception:
    pytesseract = None


def extract_text_from_source(src):
    """
    Extract text from multiple safe document formats.
    Returns empty string on unsupported or extraction failure,
    so ingestion tasks do not crash the worker.
    """

    if src.source_type in {
        "pdf", "docx", "doc", "txt", "csv", "tsv", "json", "xlsx", "svg",
        "png", "jpg", "jpeg",
    } and src.file_path:
        path = src.file_path

        # --- PDF ---
        if src.source_type == "pdf":
            reader = PdfReader(path)
            return "\n".join(page.extract_text() or "" for page in reader.pages)

        # --- Word ---
        if src.source_type in {"docx", "doc"}:
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs)

        # --- Plain text / CSV / TSV / JSON / SVG ---
        if src.source_type in {"txt", "csv", "tsv", "json", "svg"}:
            try:
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    raw = f.read()
                
                # SVG: extract text content
                if src.source_type == "svg":
                    try:
                        soup = BeautifulSoup(raw, "xml")
                        return soup.get_text(" ", strip=True)
                    except Exception:
                        return ""
                
                # JSON: pretty format
                if src.source_type == "json":
                    try:
                        obj = jsonlib.loads(raw)
                        return jsonlib.dumps(obj, ensure_ascii=False, indent=2)
                    except Exception:
                        return raw
                
                # CSV/TSV: convert to readable format
                if src.source_type in {"csv", "tsv"}:
                    delim = "\t" if src.source_type == "tsv" else ","
                    out_lines = []
                    with open(path, "r", encoding="utf-8", errors="ignore") as f2:
                        reader = csvlib.reader(f2, delimiter=delim)
                        for row in reader:
                            if row:
                                out_lines.append("\t".join(row))
                    return "\n".join(out_lines)
                
                # Plain text
                return raw
            except Exception:
                return ""

        # --- Excel ---
        if src.source_type == "xlsx":
            if load_workbook is None:
                return ""
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                lines = []
                for sheet in wb.worksheets:
                    lines.append(f"# Sheet: {sheet.title}")
                    for row in sheet.iter_rows(values_only=True):
                        if not row:
                            continue
                        vals = [("" if v is None else str(v)) for v in row]
                        if any(v.strip() for v in vals):
                            lines.append("\t".join(vals))
                return "\n".join(lines)
            except Exception:
                return ""

        # --- Images with OCR ---
        if src.source_type in {"png", "jpg", "jpeg"}:
            if Image is None or pytesseract is None:
                return ""
            try:
                img = Image.open(path)
                return pytesseract.image_to_string(img)
            except Exception:
                return ""

    # --- URL ---
    if src.source_type == "url" and src.url:
        try:
            html = requests.get(src.url, timeout=20).text
            soup = BeautifulSoup(html, "lxml")
            return soup.get_text(" ", strip=True)
        except Exception:
            return ""

    return ""


def chunk_text(text: str, max_chars=900, overlap=120):
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return []

    chunks = []
    i = 0
    while i < len(text):
        end = min(len(text), i + max_chars)
        chunk = text[i:end].strip()
        if chunk:
            chunks.append(chunk)
        i = end - overlap
        if i < 0:
            i = 0
        if end == len(text):
            break
    return chunks
